import openpyxl

wb = openpyxl.Workbook()

# Create New Sheets, Edit Title, Edit Position in WB
print ('Sheet Names: ', wb.sheetnames)
wb.create_sheet()
print ('Sheet Names: ', wb.sheetnames)
wb.create_sheet(index=0, title='First Sheet')
print ('Sheet Names: ', wb.sheetnames)
wb.create_sheet(index=2, title='Middle Sheet')
print ('Sheet Names: ', wb.sheetnames)

# Remove Sheets Based on Sheet Title
del wb['Middle Sheet']
print ('Sheet Names: ', wb.sheetnames)
del wb['Sheet1']
print ('Sheet Names: ', wb.sheetnames)